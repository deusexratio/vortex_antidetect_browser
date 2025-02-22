import os
import sys
from dataclasses import dataclass

from better_proxy import Proxy
from browserforge.fingerprints import Fingerprint
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.worksheet.worksheet import Worksheet
from loguru import logger

from db import config
from file_functions.utils import touch


@dataclass
class ProfileXLSX:
    header = ['Name', 'Proxy', 'Fingerprint (autogenerates)', 'Profile Directory (autogenerates)', 'X Token', 'Discord token']

    def __init__(self, name: str,
                 user_data_dir: str,
                 proxy: str | Proxy = '',
                 fingerprint: str | Fingerprint = '',
                 x_cookie: str | None = None,
                 discord_cookie: str | None = None,
                 ):
        self.name = name
        self.proxy = proxy
        self.fingerprint = fingerprint
        self.user_data_dir = user_data_dir
        self.x_cookie = x_cookie
        self.discord_cookie = discord_cookie


@dataclass
class WalletXLSX:
    header = ['Name', 'evm_seed_phrase', 'evm_private_key', 'solana_seed_phrase',
              'solana_private_key', 'aptos_seed_phrase', 'aptos_private_key', 'sui_seed_phrase', 'sui_private_key']

    def __init__(self,
                 name: str,
                 evm_seed_phrase: str = '', evm_private_key: str = '',
                 solana_seed_phrase: str = '', solana_private_key: str = '',
                 aptos_seed_phrase: str = '', aptos_private_key: str = '',
                 sui_seed_phrase: str = '', sui_private_key: str = '',
                 ):
        self.name = name
        self.evm_seed_phrase = evm_seed_phrase
        self.evm_private_key = evm_private_key
        self.solana_seed_phrase = solana_seed_phrase
        self.solana_private_key = solana_private_key
        self.aptos_seed_phrase = aptos_seed_phrase
        self.aptos_private_key = aptos_private_key
        self.sui_seed_phrase = sui_seed_phrase
        self.sui_private_key = sui_private_key


def create_files():
    logger.remove()
    logger.add(config.LOG_FILE, rotation="10 MB")
    logger.add(sys.stderr)

    touch(path=config.USER_FILES_DIR)
    touch(path=config.USER_DATA_DIR)
    touch(path=config.EXTENSIONS_DIR)
    touch(path=config.COOKIES_DIR)
    touch(path=config.DOWNLOADS_DIR)
    touch(path=config.X_USERNAMES_TXT, file=True)
    # touch(path=config.ADS_PROFILES_TABLE, file=True)
    # touch(path=config.IMPORT_TABLE, file=True)

    # touch(path='.env', file=True)

    def write_headers(sheet: Worksheet, headers: list[str]):
        # Стиль заголовков
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill("solid", fgColor="FFFFFF")  # Желтый фон fgColor="FFFF00"

        # Создаем стиль для границ (все стороны)
        thin = Side(border_style="thin", color="000000")  # Тонкая черная линия
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Цвета заливок
        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        green_fill = PatternFill("solid", fgColor="00FF00")
        red_fill = PatternFill("solid", fgColor="FF0000")
        purple_fill = PatternFill("solid", fgColor="800080")

        # Записываем заголовки
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = border


    if not os.path.exists(config.IMPORT_TABLE):
        workbook = Workbook()

        profiles_sheet = workbook.create_sheet(title='Profiles')
        wallets_sheet = workbook.create_sheet(title='Wallets')

        workbook.remove_sheet(worksheet=workbook['Sheet'])

        write_headers(profiles_sheet, ProfileXLSX.header)
        write_headers(wallets_sheet, WalletXLSX.header)

        workbook.save(config.IMPORT_TABLE)

    if not os.path.exists(config.ADS_PROFILES_TABLE):
        workbook = Workbook()
        sheet = workbook.active
        write_headers(sheet, ['name', 'ads id', 'cookie'])
        workbook.save(config.ADS_PROFILES_TABLE)

    if not os.path.exists(config.FANTASY_BACKUP):
        workbook = Workbook()
        sheet = workbook.active
        write_headers(sheet, ['name', 'fantasy_x_username', 'private_key', 'seed_phrase', 'fantasy_address'])
        workbook.save(config.FANTASY_BACKUP)

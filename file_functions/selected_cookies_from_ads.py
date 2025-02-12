import asyncio
import os
import sys
import time

import requests
from openpyxl.reader.excel import load_workbook
from playwright.async_api import async_playwright, Cookie
from loguru import logger

# Определяем корневую директорию проекта
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# Добавляем корневую директорию в sys.path
sys.path.append(ROOT_DIR)

from db import config
from file_functions.utils import write_json


def get_accounts_from_excel(excel_path: str) -> dict:
    profiles = {}
    workbook = load_workbook(excel_path)
    # sheet = workbook['not_done']
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
        if not row[1]:
            continue
        # print(row)
        # profile = Profile(row[0], row[1])

        name="".join(char for char in str(row[0]) if not char.isspace())
        ads_id="".join(char for char in str(row[1]) if not char.isspace())
        profiles[name] = ads_id

    logger.info(f"Получил из таблицы profiles_ads {len(profiles)} профилей")

    workbook.close()
    return profiles


async def write_cookie_to_profile_excel(excel_path: str, ads_id: str, cookie: str):
    workbook = load_workbook(excel_path)
    # sheet = workbook['not_done']
    sheet = workbook.active

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row,  min_col=1, max_col=7, values_only=True), start=2):
        if row[1] == ads_id:
            sheet.cell(row=row_num, column=3, value=cookie)

    workbook.save(excel_path)
    workbook.close()

async def write_cookie_to_profile(name: str, cookie: list[Cookie]):
    # json_path = os.path.join(config.COOKIES_DIR, f'{name}.json')
    # touch(json_path, file=True)
    write_json(path=[config.COOKIES_DIR, f'{name}.json'], obj=cookie, indent=2, encoding='utf-8')


async def open_profile(ads_id):
    args = ["--disable-popup-blocking", "--window-position=700,0"]
    args = str(args).replace("'", '"')
    open_url = f"http://local.adspower.net:50325/api/v1/browser/start?user_id=" + ads_id + f"&launch_args={str(args)}"

    try:
        # Отправка запроса на открытие профиля
        resp = requests.get(open_url).json()
        time.sleep(.5)
    except requests.exceptions.ConnectionError:
        logger.error(f'Adspower is not running.')
        sys.exit(0)
    except requests.exceptions.JSONDecodeError:
        logger.error(f'Проверьте ваше подключение. Отключите VPN/Proxy используемые напрямую.')
        sys.exit(0)
    except KeyError:
        resp = requests.get(open_url).json()

    return resp


async def task(name, ads_id, domains, semaphore, lock):
    async with semaphore:
        async with async_playwright() as p:
            close_url = "http://local.adspower.net:50325/api/v1/browser/stop?user_id=" + ads_id

            while True:
                try:
                    resp = await open_profile(ads_id)

                    browser = await p.chromium.connect_over_cdp(resp["data"]["ws"]["puppeteer"])
                    break
                except:
                    continue

            context = browser.contexts[0]
            cookies = await context.cookies(urls=domains)
            logger.debug(f"Profile {name} {ads_id} cookies: {cookies}")
            # auth_token = next((cookie['value'] for cookie in cookies
            #                    if cookie['name'] == "auth_token" and cookie['domain'] == "x.com"), None)

            # cookies_result_list = []
            # for cookie in cookies:
            #     for domain in domains:
            #         if domain in cookie['domain']:
            #             cookies_result_list.append(cookie)

                # if cookie['name'] == "auth_token":
                #     print(f"Found auth_token for {ads_id}: {cookie}")
                #     if cookie['domain'] in ["x.com", ".x.com", "twitter.com", ".twitter.com"]:
                #         auth_token = cookie['value']
                #         break

            async with lock:
                # await write_cookie_to_profile(config.ADS_PROFILES_TABLE, ads_id, str(cookies))
                await write_cookie_to_profile(name, cookies)

            requests.get(close_url)


async def main():
    semaphore = asyncio.Semaphore(3)
    lock = asyncio.Lock()

    profiles = get_accounts_from_excel(config.ADS_PROFILES_TABLE)
    domains = ['https://x.com', 'https://discord.com', 'https://live.com', 'https://twitter.com', 'https://outlook.com', 'https://hotmail.com', 'https://rambler.ru', 'https://google.com']

    tasks = [asyncio.create_task(task(name, ads_id, domains, semaphore, lock)) for name, ads_id in profiles.items()]
    await asyncio.wait(tasks)


if __name__ == '__main__':
    asyncio.run(main())

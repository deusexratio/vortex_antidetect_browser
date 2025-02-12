import json
import os.path
import random
from datetime import datetime, timedelta

from openpyxl.reader.excel import load_workbook
from better_proxy import Proxy
from loguru import logger
from browserforge.fingerprints import FingerprintGenerator, Screen
from browserforge.headers import Browser

from db import config
from db.config import USER_DATA_DIR
from db.models import db, Profile, Wallets
from db.db_api import get_profile, get_wallets_by_name
from file_functions.create_files import ProfileXLSX, WalletXLSX
from file_functions.utils import touch


def get_profiles_from_excel(excel_path: str, skip_first_line: bool = True):
    profiles = []
    workbook = load_workbook(excel_path)
    sheet = workbook["Profiles"]
    rows = sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6, values_only=True)
    for row in rows:
        if row[0]:
            name = str(row[0])
            if row[1]:
                proxy = Proxy.from_str(row[1])
            else:
                proxy = ''
            if row[2]:
                fingerprint = row[2]
            else:
                fingerprint = FingerprintGenerator(
                    browser=[
                        Browser(name='chrome', min_version=130, max_version=131),
                    ],
                    os=('windows', 'macos'),
                    device='desktop',
                    locale=('en-US',),
                    http_version=3,
                    screen=Screen(min_width=1000, min_height=600, max_width=1920, max_height=1080)
                ).generate()

            if row[3]:
                user_data_dir = str(row[3])
            else:
                user_data_dir = os.path.join(USER_DATA_DIR, name)

            if row[4] and len(row[4]) == 40:
                x_cookie = [
                    {'name': 'auth_token',
                          'value': row[4],
                          'domain': '.twitter.com', 'path': '/',
                          'expires': int((datetime.now() + timedelta(days=random.randint(300, 350))).timestamp()),
                          'httpOnly': True, 'secure': True, 'sameSite': 'None'},
                    {'name': 'auth_token',
                     'value': row[4],
                     'domain': '.x.com', 'path': '/',
                     'expires': int((datetime.now() + timedelta(days=random.randint(300, 350))).timestamp()),
                     'httpOnly': True, 'secure': True, 'sameSite': 'None'},
                            ]
            # or if you put cookie in a string it is loaded into json
            elif row[4] and len(row[4]) > 40:
                x_cookie = json.loads(row[4].replace("'", '"')
                                    .replace('True', "true")
                                    .replace('False', "false")
                                    )
            else:
                x_cookie = None

            if row[5] and len(row[5]) == 72: # todo: ПРОВЕРИТЬ ДЛИНУ
                discord_cookie = [{'name': 'token',
                          'value': row[5],
                          'domain': '.discord.com', 'path': '/',
                          'expires': int((datetime.now() + timedelta(days=random.randint(400, 450))).timestamp()),
                          'httpOnly': False, 'secure': True, 'sameSite': 'None'}]
            # or if you put cookie in a string it is loaded into json
            elif row[5] and len(row[5]) > 40:
                discord_cookie = json.loads(row[5].replace("'", '"')
                                    .replace('True', "true")
                                    .replace('False', "false")
                                    )
            else:
                discord_cookie = None

            touch(user_data_dir)

            profiles.append(
                ProfileXLSX(name=name, proxy=proxy, fingerprint=fingerprint, user_data_dir=user_data_dir,
                            x_cookie=x_cookie, discord_cookie=discord_cookie)
            )

    return profiles

def get_wallets_from_excel(excel_path: str, skip_first_line: bool = True):
    wallets = []
    workbook = load_workbook(excel_path)
    sheet = workbook["Wallets"]
    rows = sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=10, values_only=True)
    for row in rows:
        if row[0]:
            name = str(row[0])
            evm_seed_phrase = row[1]
            evm_private_key = row[2]
            solana_seed_phrase = row[3]
            solana_private_key = row[4]
            aptos_seed_phrase = row[5]
            aptos_private_key = row[6]
            sui_seed_phrase = row[7]
            sui_private_key = row[8]


            wallets.append(
                WalletXLSX(
                    name, evm_seed_phrase, evm_private_key, solana_seed_phrase, solana_private_key,
                    aptos_seed_phrase, aptos_private_key, sui_seed_phrase, sui_private_key
                           )
            )

    return wallets

def import_wallets():
    wallets = get_wallets_from_excel(config.IMPORT_TABLE)

    imported = []
    edited = []
    total = len(wallets)

    for wallet in wallets:
        wallets_instance = get_wallets_by_name(name=wallet.name)
        if wallets_instance and (
            wallets_instance.evm_seed_phrase != wallet.evm_seed_phrase or
            wallets_instance.evm_private_key != wallet.evm_private_key or
            wallets_instance.solana_seed_phrase != wallet.solana_seed_phrase or
            wallets_instance.solana_private_key != wallet.solana_private_key or
            wallets_instance.aptos_seed_phrase != wallet.aptos_seed_phrase or
            wallets_instance.aptos_private_key != wallet.aptos_private_key or
            wallets_instance.sui_seed_phrase != wallet.sui_seed_phrase or
            wallets_instance.sui_private_key != wallet.sui_private_key
        ):
            wallets_instance.evm_seed_phrase = wallet.evm_seed_phrase
            wallets_instance.evm_private_key = wallet.evm_private_key
            wallets_instance.solana_seed_phrase = wallet.solana_seed_phrase
            wallets_instance.solana_private_key = wallet.solana_private_key
            wallets_instance.aptos_seed_phrase = wallet.aptos_seed_phrase
            wallets_instance.aptos_private_key = wallet.aptos_private_key
            wallets_instance.sui_seed_phrase = wallet.sui_seed_phrase
            wallets_instance.sui_private_key = wallet.sui_private_key

            db.commit()
            edited.append(wallets_instance)

        if not wallets_instance:
            wallets_instance = Wallets(
                name=wallet.name,
                evm_seed_phrase = wallet.evm_seed_phrase,
                evm_private_key = wallet.evm_private_key,
                solana_seed_phrase = wallet.solana_seed_phrase,
                solana_private_key = wallet.solana_private_key,
                aptos_seed_phrase = wallet.aptos_seed_phrase,
                aptos_private_key = wallet.aptos_private_key,
                sui_seed_phrase = wallet.sui_seed_phrase,
                sui_private_key = wallet.sui_private_key
            )

            db.insert(wallets_instance)
            imported.append(wallets_instance)

    logger.success(f'Done! imported wallets for profiles: {len(imported)}/{total}; '
                       f'edited: {len(edited)}/{total}; total: {total}')

def import_profiles():
    profiles = get_profiles_from_excel(config.IMPORT_TABLE)

    imported = []
    edited = []
    total = len(profiles)

    for profile in profiles:
        profile_instance = get_profile(name=profile.name)

        if profile_instance and profile_instance.proxy != profile.proxy.as_url:
            print(f"{profile.name}, old proxy {profile_instance.proxy}, new proxy {profile.proxy.as_url}")
            profile_instance.proxy = profile.proxy.as_url

            db.commit()
            edited.append(profile_instance)

        if not profile_instance:
            profile_instance = Profile(
                name=profile.name,
                proxy=profile.proxy.as_url,
                page_urls=[''],
                fingerprint=profile.fingerprint.dumps(),
                user_data_dir=profile.user_data_dir,
                # last_opening_time=None,
            )
            db.insert(profile_instance)
            imported.append(profile_instance)
            logger.debug(profile_instance)

    logger.success(f'Done! imported profiles: {len(imported)}/{total}; '
                   f'edited profiles: {len(edited)}/{total}; total: {total}')

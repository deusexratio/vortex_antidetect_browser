import os
import sys
import asyncio

from openpyxl import load_workbook
from openpyxl.styles import Side, Border
from playwright.async_api import async_playwright, BrowserContext
from loguru import logger

from file_functions.utils import get_file_names

# Определяем корневую директорию проекта
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# Добавляем корневую директорию в sys.path
sys.path.append(ROOT_DIR)

from browser_functions.functions import launch_profile_async
from db.db_api import get_profile
from db.models import Profile
from db import config


class Fantasy:
    reg_url = 'https://monad.fantasy.top?ref='
    base_url = 'https://monad.fantasy.top'
    login_url = 'https://monad.fantasy.top/login'
    shop_url = 'https://monad.fantasy.top/shop'
    # Добавляем классовую переменную для хранения использования кодов
    ref_code_usage = {}

    @classmethod
    def initialize_ref_codes(cls, ref_codes):
        """Инициализация счетчиков для реферральных кодов"""
        cls.ref_code_usage = {code: 0 for code in ref_codes}

    def __init__(self, context: BrowserContext, profile: Profile, ref_codes: list[str] | None, reg_for_ref_code: int):
        self.context = context
        self.profile = profile
        self.ref_codes = ref_codes
        self.reg_for_ref_code = reg_for_ref_code

    async def reg_acc(self):
        fantasy_page = await self.context.new_page()

        try:
            await fantasy_page.goto(self.login_url)
        except Exception as e:
            if 'ERR_HTTP_RESPONSE_CODE_FAILURE' in str(e):
                logger.error(
                    f"Profile {self.profile.name} | Proxy authentication error (407). Please check proxy credentials.")
                await asyncio.sleep(10)
                try:
                    await fantasy_page.goto(self.login_url)
                except Exception as e:
                    raise e

        await asyncio.sleep(3)

        register_button = fantasy_page.locator('//button').first
        logger.debug(f"Profile {self.profile.name} | Clicking on Register button")
        await register_button.click(timeout=10000)

        await asyncio.sleep(3)
        if await fantasy_page.get_by_text('Checking that').is_visible(timeout=15000):
            if await fantasy_page.get_by_text('CAPTCHA passed successfully.').is_visible(timeout=15000):
                logger.success(f"Profile {self.profile.name} | Passed CAPTCHA, clicking Continue")
                await fantasy_page.get_by_text('Continue', exact=True).click(timeout=10000)
            else:
                logger.warning(f"Profile {self.profile.name} | Didn't pass CAPTCHA")

        logger.debug(f"Profile {self.profile.name} | Clicking on Twitter button")
        await fantasy_page.get_by_text('Twitter', exact=True).click(timeout=10000)

        await asyncio.sleep(5)

        logger.debug(f"Profile {self.profile.name} | Clicking on Authorize button")
        await fantasy_page.get_by_test_id('OAuth_Consent_Button').click(timeout=10000)
        await asyncio.sleep(15)
        if await fantasy_page.get_by_text('Your Twitter account is suspended.').is_visible(timeout=15000):
            logger.error(f"Profile {self.profile.name} | Your Twitter account is suspended.")

        logger.debug(f"Profile {self.profile.name} | Filling ref code")
        # Находим код, который использовался меньше reg_for_ref_code раз
        available_codes = [code for code, usage in self.__class__.ref_code_usage.items()
                         if usage < self.reg_for_ref_code]

        if not available_codes:
            logger.error(f"Profile {self.profile.name} | All referral codes have reached their limit of {self.reg_for_ref_code} uses")
            return False

        # Выбираем код с наименьшим количеством использований среди доступных
        current_ref = min(
            ((code, self.__class__.ref_code_usage[code]) for code in available_codes),
            key=lambda x: x[1]
        )[0]

        self.__class__.ref_code_usage[current_ref] += 1
        logger.debug(f"Profile {self.profile.name} | Using ref code {current_ref} (used {self.__class__.ref_code_usage[current_ref]} times)")
        await fantasy_page.get_by_placeholder('Add your Referal Code').fill(current_ref)

        for i in range(5):
            logger.debug(f"Profile {self.profile.name} | Clicking on Continue for {i} time")
            await fantasy_page.get_by_text('Continue', exact=True).click(timeout=10000)
            await asyncio.sleep(.6)
        # await fantasy_page.get_by_text('Continue', exact=True).click(timeout=10000)
        # await fantasy_page.get_by_text('Continue', exact=True).click(timeout=10000)
        # await fantasy_page.get_by_text('Continue', exact=True).click(timeout=10000)
        # await fantasy_page.get_by_text('Continue', exact=True).click(timeout=10000)

        logger.debug(f"Profile {self.profile.name} | Clicking on Claim 100 $fMON")
        await fantasy_page.get_by_text('Claim 100 $fMON', exact=True).click(timeout=10000)
        await asyncio.sleep(10)

        while True:
            logger.debug(f"Profile {self.profile.name} | Waiting for faucet")
            if not await fantasy_page.get_by_text('deploying more faucet - steady lads').is_visible(timeout=5000):
                break
            await asyncio.sleep(3)

        try:
            await fantasy_page.get_by_text('Close', exact=True).click(timeout=10000)
        except Exception as e:
            logger.error(f"Profile {self.profile.name} | {e}")

        logger.success(f"Fully registered profile {self.profile.name}")
        return True

    async def spin_roulette(self):
        shop_page = await self.context.new_page()
        await shop_page.goto(self.shop_url)
        for i in range(5):
            logger.debug(f"Profile {self.profile.name} | Clicking on Continue for {i} time")
            await shop_page.get_by_text('Continue', exact=True).click(timeout=10000)
            await asyncio.sleep(.6)

        logger.debug(f"Profile {self.profile.name} | Clicking on Got it, let's go")
        await shop_page.get_by_text("Got it, let's go", exact=True).click(timeout=10000)

        logger.debug(f"Profile {self.profile.name} | Clicking on Claim")
        await shop_page.get_by_text("Claim", exact=True).click(timeout=10000)
        await asyncio.sleep(7)

        logger.debug(f"Profile {self.profile.name} | Clicking on Confirm")
        await shop_page.get_by_text("Confirm", exact=True).click(timeout=10000)
        return True

    async def export_backup(self):
        fantasy_page = await self.context.new_page()
        await fantasy_page.goto(self.base_url)
        await asyncio.sleep(5)

        username_selector = fantasy_page.locator('//*[@id="sidebar"]/div[3]/div[5]/a/div/div[1]/div/h2/span')
        username = await username_selector.text_content()

        await username_selector.click()
        await fantasy_page.get_by_text("Settings", exact=True).click()
        await fantasy_page.get_by_text("Export Wallet", exact=True).click()
        await asyncio.sleep(5)

        logger.debug(f"Profile {self.profile.name} | Getting private key")
        # Получаем private key
        copy_key_button = fantasy_page.locator("//button").nth(21)
        private_key = await fantasy_page.evaluate("""() => {
            return new Promise((resolve) => {
                const listener = (e) => {
                    document.removeEventListener('copy', listener);
                    resolve(e.clipboardData.getData('text/plain'));
                };
                document.addEventListener('copy', listener);
                document.execCommand('copy');
            });
        }""", copy_key_button)
        await copy_key_button.click()
        await asyncio.sleep(1)
        logger.success(f"Profile {self.profile.name} | private key {private_key}")

        logger.debug(f"Profile {self.profile.name} | Getting seed phrase")
        # Получаем seed phrase
        copy_phrase_button = fantasy_page.locator("//button").nth(22)
        seed_phrase = await fantasy_page.evaluate("""() => {
            return new Promise((resolve) => {
                const listener = (e) => {
                    document.removeEventListener('copy', listener);
                    resolve(e.clipboardData.getData('text/plain'));
                };
                document.addEventListener('copy', listener);
                document.execCommand('copy');
            });
        }""", copy_phrase_button)
        await copy_phrase_button.click()
        await asyncio.sleep(1)
        logger.success(f"Profile {self.profile.name} | seed phrase {seed_phrase}")

        fantasy_address = fantasy_page.url.split('/')[-1]
        write_backup_for_profile(config.FANTASY_BACKUP, self.profile, username, private_key, seed_phrase, fantasy_address)
        return True


async def fantasy_launcher(semaphore, profile, ref_code, reg_for_ref_code, do_register, do_spin, do_backup):
    async with semaphore:
        async with async_playwright() as playwright_instance:
            try:
                context = await launch_profile_async(profile=profile, playwright_instance=playwright_instance,
                                                     extensions=extensions,
                                                     keep_open=False, restore_pages=True)

                fantasy = Fantasy(context, profile, ref_code, reg_for_ref_code)

                if do_register:
                    result_reg = await fantasy.reg_acc()
                    if result_reg is False:  # Если все коды достигли лимита или произошла ошибка
                        logger.error(f"Profile {profile.name} registration skipped - all ref codes reached limit!")
                    if result_reg:
                        logger.success(f"Profile {profile.name} registered account in Fantasy Monad!")

                if do_spin:
                    result_spin = await fantasy.spin_roulette()
                    if result_spin:
                        logger.success(f"Profile {profile.name} spinned daily roulette!")

                if do_backup:
                    result_backup = await fantasy.export_backup()
                    if result_backup:
                        logger.success(f"Profile {profile.name} exported backup!")

                Results.success += 1

                # Закрываем контекст после выполнения всех операций
                await context.close()

            except Exception as e:
                error_msg = str(e)
                if "407" in error_msg:
                    logger.error(f"Profile {profile.name} | Proxy authentication error (407). Please check proxy credentials.")

                logger.error(f"Profile {profile.name} error {e}")
                await asyncio.sleep(5)
                Results.errors += 1
                logger.warning(f"Profile {profile.name} failed")


extensions = get_file_names(config.EXTENSIONS_DIR, files=False)


class Results:
    @classmethod
    def results(cls):
        cls.success = 0
        cls.errors = 0


def write_backup_for_profile(excel_path: str, profile: Profile, username: str,
                                  backup_pk: str, backup_seed: str, fantasy_address: str):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # Создаем стиль для границ (все стороны)
    thin = Side(border_style="thin", color="000000")  # Тонкая черная линия
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Определяем номер строки для записи (следующая свободная строка)
    row_num = sheet.max_row + 1

    # Записываем данные в новую строку
    sheet.cell(row=row_num, column=1, value=profile.name).border = border
    sheet.cell(row=row_num, column=2, value=username).border = border
    sheet.cell(row=row_num, column=3, value=backup_pk).border = border
    sheet.cell(row=row_num, column=4, value=backup_seed).border = border
    sheet.cell(row=row_num, column=4, value=fantasy_address).border = border

    workbook.save(excel_path)
    workbook.close()


async def main():
    if len(sys.argv) != 8:
        print("Usage: python fantasy_monad.py <profiles> <ref_codes> <reg_for_ref_code> <threads> <do_register> <do_spin> <do_backup>")
        sys.exit(1)

    profiles_list_str = sys.argv[1]
    ref_codes = sys.argv[2]
    reg_for_ref_code = int(sys.argv[3])
    threads = int(sys.argv[4])
    do_register = bool(int(sys.argv[5]))
    do_spin = bool(int(sys.argv[6]))
    do_backup = bool(int(sys.argv[7]))

    semaphore = asyncio.Semaphore(threads)

    ref_codes = ref_codes.split(',') if ref_codes else []
    Fantasy.initialize_ref_codes(ref_codes)
    Results.results()

    profiles = [get_profile(profile_name) for profile_name in profiles_list_str.split(',')]

    tasks = [asyncio.create_task(fantasy_launcher(semaphore, profile, ref_codes, reg_for_ref_code,
                                                do_register, do_spin, do_backup)) for profile in profiles]
    logger.debug(f"Launching {len(tasks)} tasks")
    await asyncio.wait(tasks)
    logger.success(f"Executed {Results.success} successful tasks and failed {Results.errors} tasks")
    sys.exit(0)

if __name__ == "__main__":
    asyncio.run(main())
